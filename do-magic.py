import os
import warnings
from tqdm import tqdm
import string
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from openpyxl.styles import Font

warnings.simplefilter("ignore")

def set_border(ws, cell_range):
    thin = openpyxl.styles.Side(border_style="thin")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
            cell.comment = None
            # cell.alignment = Alignment(wrap_text=True)


def set_header(ws, cell_range, color):
    ws.row_dimensions[1].height = 30
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = openpyxl.styles.PatternFill(start_color="ffffff", fill_type="solid")
            cell.font = cell.font.copy(color=color)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

def center_range(ws, cell_range):
    ws.row_dimensions[1].height = 30
    for row in ws[cell_range]:
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center', wrap_text=True)

def set_header_font_size_14(ws, cell_range):
    ws.row_dimensions[1].height = 30
    for row in ws[cell_range]:
        for cell in row:
            cell.fill = openpyxl.styles.PatternFill(start_color="ffffff", fill_type="solid")
            cell.font = cell.font.copy(color="000000", size = "14")
            cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

def check_end(ws, start, col):
    no_end = True
    start = 8
    while no_end:
        if ws['{}{}'.format(col, start)].value:
            pass
        else:
            no_end = False
            return start - 1
        start += 1


def check_max_col(ws, start_row):
    no_end = True
    i = 8
    while no_end:
        if not ws['{}{}'.format(get_column_letter(i), start_row)].value:
            no_end = False
        i += 1
    return get_column_letter(i - 2)


def check_start_a(ws, start, end):
    for i in range(start, end):
        if ws['a{}'.format(i)].value and not (ws['a{}'.format(i)].value) == '.':
            return (i)


def check_start_f(ws):
    col_range = list(string.ascii_lowercase)
    for i, n in enumerate(col_range):
        if ws['{}2'.format(n)].value  and not (ws['{}2'.format(n)].value) == '.':
            return col_range[i], col_range[i - 1], col_range[i + 1]

def check_for_hide_colums(n, e, ws):
    to_hide = []
    col_range = list(string.ascii_lowercase)
    for i, col in enumerate(col_range):
        if '[Attr' in str(ws['{}{}'.format(col,n)].value):
            to_hide.append(col)
    return(to_hide)


def hide_cols(to_hide, ws):
    for col in to_hide:
        ws.column_dimensions[col.upper()].hidden = True       
    return ws


def check_for_hide_rows(n, m, ws):
    to_hide = []
    for row in range(1, n):
        if '[Attr' in str(ws['{}{}'.format(m, row)].value):
            to_hide.append(row)
    return to_hide


def hide_rows(to_hide, ws):
    for row in to_hide: 
        ws.row_dimensions[row].hidden = True     
    return ws    


def format_first_type(ws):
    # default_column_width = 25
    # # for i in list(string.ascii_lowercase):
    # #     ws.column_dimensions[i].width = default_column_width
    
    n = check_start_a(ws, 5 , 19) # start of first block
    m, e , x = check_start_f(ws) # m = start column of second block; e = end column of first block; x = freeze column
    end = check_end(ws, 8, 'A') # end of first block
    col = check_max_col(ws, 5) # end column
    to_hide = check_for_hide_colums(n, e, ws)
    ws = hide_cols(to_hide, ws)
    to_hide = check_for_hide_rows(n, m, ws)
    ws = hide_rows(to_hide, ws)    
    set_border(ws, 'A{}:{}{}'.format(n, e, end))
    set_border(ws, '{}1:{}{}'.format(m, col, end))
    set_header(ws, 'A{}:{}{}'.format(n, e, n), '000000')
    set_header(ws, '{}1:{}{}'.format(m, m, n - 1), '000000')
    ws['A1'].alignment = Alignment(horizontal='left', vertical = 'top', wrap_text=True)
    ws['A1'].font = Font(size="9", bold=True)
    ws.merge_cells('A1:{}{}'.format(e, n -1))
    ws.freeze_panes = '{}{}'.format(x, n+1)
    ws.row_dimensions[1].height = 30
    return ws


def format_transformation_table(ws):
    default_column_width = 25
    for i in list(string.ascii_lowercase):
        ws.column_dimensions[i].width = default_column_width
    n = check_start_a(ws, 2 , 19) # start of first block
    m, e , x = check_start_f(ws) # m = start column of second block; e = end column of first block; x = freeze column
    end = check_end(ws, 8, 'A') # end of first block
    col = check_max_col(ws, 2) # end column
    # to_hide = check_for_hide_colums(n, e, ws)
    # ws = hide_cols(to_hide, ws)
    # to_hide = check_for_hide_rows(n, m, ws)
    # ws = hide_rows(to_hide, ws)    
    
    set_border(ws, 'A{}:{}{}'.format(n, e, end))
    set_border(ws, '{}1:{}{}'.format(m, col, end))
    set_header(ws, 'A{}:{}{}'.format(n, e, n), '000000')
    set_header(ws, '{}1:{}{}'.format(m, m, n - 1), '000000')
    ws['A1'].alignment = Alignment(horizontal='left', vertical = 'top', wrap_text=True)
    ws['A1'].font = Font(size="9", bold=True)
    ws.merge_cells('A1:{}{}'.format(e, n -1))
    ws.freeze_panes = '{}{}'.format(x, n+1)
    ws.row_dimensions[1].height = 30
    return ws


def next_alpha(s):
    return chr((ord(s.upper())+1 - 65) % 26 + 65)


def format_information_result(ws, last_tab):
    default_column_width = 20
    end = check_end(ws, 1, 'B') + 1
    
    for i in range(1, end):
        if ws['a1'].value == "ID":
            ws['a{}'.format(i)].alignment = Alignment(horizontal='center')
        ws['{}{}'.format(last_tab, i)].alignment = Alignment(horizontal='center')     

    for i in list(string.ascii_lowercase):
        ws.column_dimensions[i].width = default_column_width
        if ws['a1'].value == "ID" and i == "a":
            ws.column_dimensions[i].width = "5"
        if ws['{}1'.format(i)].value == "Status" :
            ws.column_dimensions[i].width = "40"
        if ws['{}1'.format(i)].value == "Condition" :
            ws.column_dimensions[i].width = "28"
        elif ws['{}1'.format(i)].value == "Process" or ws['{}1'.format(i)].value == "C/E":
            ws.column_dimensions[i].width = "10"
            for row in range(1, end):
                ws['{}{}'.format(i, row)].alignment = Alignment(horizontal='center')

    ws.column_dimensions[last_tab].width = 5
    ws.merge_cells('{}1:{}1'.format(last_tab, next_alpha(last_tab)))
    set_header_font_size_14(ws, 'A1:{}1'.format(last_tab))
    ws.freeze_panes = ws['a2']
    return ws

def format_status_table(ws, last_tab):
    default_column_width = 37
    # ws.merge_cells('b1:c1')
    # ws.merge_cells('e1:f1')
    try:
        if str(ws['b1'].value) == 'None':
            ws['b1'].value = ws['a1'].value
            ws['b1'].font = Font(bold=True, name='Dialog.bold')
            ws['a1'].value = ""
    except AttributeError:
        pass
        
    set_header_font_size_14(ws, 'A1:{}2'.format(last_tab))
    ws.freeze_panes = ws['a3']
    for i in list(string.ascii_lowercase):
        ws.column_dimensions[i].width = default_column_width
        if i == "a" : ws.column_dimensions[i].width = "13"

        if ("IDENTIFIER" in str(ws['{}2'.format(i)].value) or ws['{}1'.format(i)].value == "to") and (str(ws['{}3'.format(i)].value).isdigit() or str(ws['{}3'.format(i)].value) == 'None') :
            ws.column_dimensions[i].width = "6"
            ws['{}2'.format(i)].alignment = Alignment(horizontal='left')
            ws['{}1'.format(i)].alignment = Alignment(horizontal='left')
    return ws


def find_last_tab_2(ws):
    col_range = list(string.ascii_lowercase)
    for c in col_range:
        if "Document Status" in str(ws['{}1'.format(c)].value) and c != "a" :
            return next_alpha(c)
    return "f"


def find_last_tab(ws):
    col_range = list(string.ascii_lowercase)
    for c in col_range:
        if  "CERTEX" in str(ws['{}1'.format(c)].value):
            return c
    return "d"

def column_letters():
    col_range = list(string.ascii_lowercase)
    new_list = list(string.ascii_lowercase)

    for c in col_range:
        for letter in col_range:
            new_list.append(c + letter)
    return new_list


def format_information_result_recap(ws):
    ws.freeze_panes = ws['h3']
    ws.sheet_view.zoomScale = 70
    center_range(ws, 'a1:g30')

    for col, height in [[1, 30], [2, 51]]:
        ws.row_dimensions[col].height = height

    for col, width in [['B', 43], ['G', 14], ['N', 29], ['AG', 29]]:
        ws.column_dimensions[col].width = width

    for column in ['A', 'H', 'K', 'R', 'U', 'X', 'AA', 'AD', 'AF']:
        ws.column_dimensions[column].width = 5

    for column in ['C', 'D', 'E', 'F']:
        ws.column_dimensions[column].width = 17

    for column in ['J', 'M', 'O', 'Q', 'T', 'W', 'Z', 'AC', 'AE', 'AH']:
        ws.column_dimensions[column].width = 26

    for column in ['I', 'L', 'S', 'V', 'Y', 'AB', 'AD']:
        ws.column_dimensions[column].width = 40

    for range in ['A1:B1', 'C1:F1', 'H1:J1', 'K1:M1', 'P1:Q1', 'R1:T1', 'U1:W1', 'X1:Z1', 'AA1:AC1', 'AD1:AE1', 'AF1:AH1']:
        ws.merge_cells(range)

    return ws

def main():
    done_folder = "done"
    work_folder = "work"
    full_work_folder = os.path.join(os.path.dirname(os.path.realpath(__file__)), work_folder, "")
    dir_list = []

    for file in  os.listdir(full_work_folder):
        if file.endswith(".xlsx"):
            dir_list.append(file)

    print("Formatting all {} .xlsx files found in {}".format(len(dir_list), full_work_folder))   

    for filename in tqdm(dir_list, desc ="Work done: "):
        if "Transformation Table" in filename and "Status" not in filename and "to" in filename:
            wb = openpyxl.load_workbook(os.path.join(full_work_folder, filename))
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                ws.sheet_view.zoomScale = 70
                ws = format_first_type(ws)

            wb.save(os.path.join(os.path.dirname(os.path.realpath(__file__)), done_folder, filename))

        elif "Transformation Table" in filename and "Status" not in filename and "to" not in filename:
            wb = openpyxl.load_workbook(os.path.join(full_work_folder, filename))
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                ws.sheet_view.zoomScale = 70
                ws = format_transformation_table(ws)
            
            wb.save(os.path.join(os.path.dirname(os.path.realpath(__file__)), done_folder, filename))      

        elif "Information Result" in filename:
            wb = openpyxl.load_workbook(os.path.join(full_work_folder, filename))
            for ws_name in wb.sheetnames:
                if "Recap" not in ws_name:
                    ws = wb[ws_name]
                    ws.sheet_view.zoomScale = 70
                    ws = format_information_result(ws, find_last_tab(ws))                   
                else:
                    ws = wb[ws_name]
                    ws = format_information_result_recap(ws)
            wb.save(os.path.join(os.path.dirname(os.path.realpath(__file__)), done_folder, filename))
        
        elif "Status Transformation Table" in filename:
            wb = openpyxl.load_workbook(os.path.join(full_work_folder, filename))
            for ws_name in wb.sheetnames:

                ws = wb[ws_name]
                ws.sheet_view.zoomScale = 70
                ws = format_status_table(ws, find_last_tab_2(ws))
            
            set_header(wb['CERTEX Statuses'], 'C1:F2', '999999')
            wb.save(os.path.join(os.path.dirname(os.path.realpath(__file__)), done_folder, filename))

        elif "Spreadsheet Rules Table" in filename:
            wb = openpyxl.load_workbook(os.path.join(full_work_folder, filename))
            col_range = column_letters()
            
            for sheet, cell in [[0, 'D6'], [1, 'C6'], [2, 'C6'], [3, 'C5'], [4, 'F6'], [5, 'C6'], [6, 'C6'], [7, 'D6'], [8, 'C6']]:
                try:
                    wb[wb.sheetnames[sheet]].freeze_panes = wb[wb.sheetnames[sheet]][cell]
                except:
                    pass

            for sheet, col, width in [[0, 'B', 63.44], [4, 'B', 36], [4, 'C', 39], [4, 'D', 43], [5, 'A', 35], [6, 'A', 15], [7, 'A', 15], [8, 'A', 15]]:
                try:
                    wb[wb.sheetnames[sheet]].column_dimensions[col].width = width
                except:
                    pass

            for sheet, cell in [[1, 'A6'], [2, 'A5'], [3, 'A5'], [5, 'A5'], [6, 'B5'], [7, 'B5'], [8, 'A5']]:
                try:
                    wb[wb.sheetnames[sheet]][cell].font = Font(bold=True, name='Dialog.bold', size=12)
                except:
                    pass
                

            for sheet, col_source, col_dest in [[3, 'A5', 'A1'], [6, 'B5', 'B1'], [7, 'B5', 'B1']]:
                try:
                    wb[wb.sheetnames[sheet]][col_source].value = wb[wb.sheetnames[sheet]][col_dest].value
                    wb[wb.sheetnames[sheet]][col_dest].value = ""
                except:
                    pass

            wb[wb.sheetnames[3]]['a5'].alignment = Alignment(horizontal='left')

            for name in wb.sheetnames:
                ws = wb[name]
                ws.sheet_view.zoomScale = 70
                ws.row_dimensions[1].height = 30 
                ws.row_dimensions[2].height = 180
                ws.row_dimensions[4].height = 180
                         
                for column in col_range:
                    ws['{}1'.format(column)].font = Font(size = '14', bold = True, name='Dialog.bold')
                    ws['{}2'.format(column)].font = Font(size = '10', name='Dialog.plain')
                    ws['{}4'.format(column)].font = Font(size = '10', name='Dialog.plain')
                    ws['{}2'.format(column)].alignment = Alignment(textRotation = 90, horizontal='left', wrap_text = True)
                    ws['{}4'.format(column)].alignment = Alignment(textRotation = 90, horizontal='left', wrap_text = True)

            wb.save(os.path.join(os.path.dirname(os.path.realpath(__file__)), done_folder, filename))
        elif "Spreadsheet Record Outcome" in filename:
            wb = openpyxl.load_workbook(os.path.join(full_work_folder, filename))
            # for name in wb.sheetnames:
            ws = wb['Recap']
            ws.sheet_view.zoomScale = 70
            # ws.delete_rows(1, 1)
            center_range(ws, 'A2:H3')
            for column in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws[f'{column}2'].font = Font(bold=True, name='Dialog.bold')
                ws[f'{column}3'].font = Font(bold=True, name='Dialog.bold')
                
            for range in ['A2:A3', 'B2:B3', 'G2:G3', 'H2:H3']:
                ws.merge_cells(range)              
                 
            wb.save(os.path.join(os.path.dirname(os.path.realpath(__file__)), done_folder, filename))
        elif "Spreadsheet Measurement Units Transformation" in filename:
            wb = openpyxl.load_workbook(os.path.join(full_work_folder, filename))
            for name in wb.sheetnames:
                ws = wb[name]
                ws.sheet_view.zoomScale = 70

            wb.save(os.path.join(os.path.dirname(os.path.realpath(__file__)), done_folder, filename))       
        else:
            wb = openpyxl.load_workbook(os.path.join(full_work_folder, filename))
            for name in wb.sheetnames:
                ws = wb[name]
                ws.sheet_view.zoomScale = 70

            wb.save(os.path.join(os.path.dirname(os.path.realpath(__file__)), done_folder, filename))             
            print(f'{filename} is not supported')
        
 
if __name__ == '__main__':
    main()